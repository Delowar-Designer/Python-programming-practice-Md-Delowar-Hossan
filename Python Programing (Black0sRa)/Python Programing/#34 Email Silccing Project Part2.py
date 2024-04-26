# #34 Email Silccing Project Part2
from tkinter import *
from openpyxl import *

wb = load_workbook('EmailSlice.xlsx')
sheet = wb.active


def excel_file():
    sheet.column_dimensions['A'].wiidth = 30
    sheet.column_dimensions['B'].wiidth = 30
    sheet.column_dimensions['C'].wiidth = 30

    sheet.cell(row=1, column=1).value = 'E-mail Name'
    sheet.cell(row=1, column=2).value = 'User Name'
    sheet.cell(row=1, column=3).value = 'Domain Name'


def clear_text():
    text_field.delete(0, END)


def clickme():
    temp_text = text_field.get()
    user_name = temp_text[:temp_text.index('@')]
    domain_name = temp_text[temp_text.index('@') + 1:]

    result = f"E-mail: {temp_text}\n" \
             f"User Name: {user_name}\n" \
             f"Domain Name: {domain_name}\n"

    output.insert(END, result)
    # text_field.delete(0, END)

    if text_field.get() == "":
        print("Empty Line")
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = text_field.get()
        sheet.cell(row=current_row + 1, column=2).value = user_name
        sheet.cell(row=current_row + 1, column=3).value = domain_name

        wb.save('EmailSlice3.xlsx')
        clear_text()


# Drivr Code
if __name__ == '__main__':
    window = Tk()
    window.title("Email Slicing Programer Delowar.Designer")
    window.geometry("800x500")

    excel_file()

    # toplevel frame
    frame1 = Label(window, width=500, height=2, text="E-mail Slicing", font=("Bold", 20), bg='black', fg='white')
    frame1.pack()

    # top text email writing
    label1 = Label(window, text="Write an E-mail", font=('Arial', 15), fg='#010847')
    label1.pack(pady=15)

    email_name = StringVar()

    # user input box
    text_field = Entry(window, text = email_name, font=('Arial', 20), bg='#8dc94d', width=25)
    text_field.pack()

    # slice button
    btn1 = Button(window, text='Click here to Slice', font=("Bold", 15), bg='#8591de', command=clickme)
    btn1.pack(pady=40)

    # output box
    output = Text(window, width=100, height=15, bg='#c195e6', fg='black', font=('Arial', 15))
    output.pack()

    # exit Button
    btn2 = Button(window, text="Exit Button", font=("Bold", 10), bg='red', fg='white', command=window.destroy)
    btn2.pack(pady=20)

    window.mainloop()
